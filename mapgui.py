"""
wro_gui.py  —  WRO 2026 Mosaic Masters
=======================================
Pon mapa_wro2026.jpg en la misma carpeta que este script.

pip install bleak openpyxl Pillow
pip install winrt-Windows.Devices.Bluetooth.Advertisement
pip install winrt-Windows.Storage.Streams
"""

import asyncio, math, os, struct, threading, time
import tkinter as tk
from tkinter import filedialog, messagebox, colorchooser
from collections import deque

from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from bleak import BleakScanner

# ════════════════════════════════════════════════════════════════
#  PROTOCOLO PYBRICKS
# ════════════════════════════════════════════════════════════════
PYBRICKS_MFR_ID = 0x0397
TX_CHANNEL = 1
RX_CHANNEL = 0
TYPE_INT   = 3
TURN_RATE  = 200
MOTOR_SPD  = 300

def _enc(v: int) -> bytes:
    v = int(v)
    if -128 <= v <= 127:
        return bytes([(TYPE_INT << 5) | 1]) + struct.pack("b", v)
    elif -32768 <= v <= 32767:
        return bytes([(TYPE_INT << 5) | 2]) + struct.pack("<h", v)
    else:
        return bytes([(TYPE_INT << 5) | 4]) + struct.pack("<i", v)

def encode_cmd(speed, turn, motor_cmd=0, motor_val=0, turn_cmd=0) -> bytes:
    return (bytes([TX_CHANNEL])
            + _enc(speed) + _enc(turn)
            + _enc(motor_cmd) + _enc(motor_val)
            + _enc(turn_cmd))

def decode_hub(data: bytes):
    """Decodifica (dist_mm, heading, r, g, b)."""
    if not data or data[0] != RX_CHANNEL:
        return None
    values, i = [], 1
    while i < len(data):
        header = data[i]; i += 1
        vt = (header >> 5) & 0x07
        vl =  header & 0x1F
        if vt == 3:
            if i + vl > len(data): break
            raw = data[i:i+vl]
            v = (struct.unpack("b", raw)[0] if vl == 1
                 else struct.unpack("<h", raw)[0] if vl == 2
                 else struct.unpack("<i", raw)[0])
            values.append(int(v)); i += vl
        elif vt == 4:
            if i + 4 > len(data): break
            values.append(float(struct.unpack_from("<f", data, i)[0])); i += 4
        else:
            i += vl
    return values if len(values) >= 5 else None

# ════════════════════════════════════════════════════════════════
#  BLE PUBLISHER — hilo dedicado, bajo delay
# ════════════════════════════════════════════════════════════════
class BLEPublisher:
    REFRESH_S = 2.0

    def __init__(self):
        self._speed = self._turn = self._mc = self._mv = self._tc = 0
        self._changed  = threading.Event()
        self._stop_evt = threading.Event()
        threading.Thread(target=self._loop, daemon=True).start()

    def _ibuf(self, data):
        from winrt.windows.storage.streams import DataWriter
        w = DataWriter()
        for b in data: w.write_byte(b)
        return w.detach_buffer()

    def _make_pub(self, s, t, mc, mv, tc):
        from winrt.windows.devices.bluetooth.advertisement import (
            BluetoothLEAdvertisementPublisher,
            BluetoothLEAdvertisement,
            BluetoothLEManufacturerData,
        )
        mfr = BluetoothLEManufacturerData()
        mfr.company_id = PYBRICKS_MFR_ID
        mfr.data = self._ibuf(encode_cmd(s, t, mc, mv, tc))
        adv = BluetoothLEAdvertisement()
        adv.manufacturer_data.append(mfr)
        pub = BluetoothLEAdvertisementPublisher(adv)
        pub.start()
        return pub

    def _loop(self):
        pub = None; cur = (None,)*5; last_ref = 0.0
        while not self._stop_evt.is_set():
            self._changed.wait(timeout=self.REFRESH_S)
            self._changed.clear()
            if self._stop_evt.is_set(): break
            nxt = (self._speed, self._turn, self._mc, self._mv, self._tc)
            now = time.monotonic()
            if nxt != cur or (now - last_ref) >= self.REFRESH_S or pub is None:
                if pub:
                    try: pub.stop()
                    except: pass
                try:
                    pub = self._make_pub(*nxt); cur = nxt; last_ref = now
                except Exception as e:
                    print(f"[BLE TX] {e}"); pub = None
        if pub:
            try: pub.stop()
            except: pass

    def send(self, speed=0, turn=0, mc=0, mv=0, tc=0):
        changed = (speed,turn,mc,mv,tc) != (self._speed,self._turn,self._mc,self._mv,self._tc)
        self._speed=speed; self._turn=turn; self._mc=mc; self._mv=mv; self._tc=tc
        if changed: self._changed.set()

    def stop(self):
        self._stop_evt.set(); self._changed.set()

# ════════════════════════════════════════════════════════════════
#  POSICIÓN REAL DEL HUB
# ════════════════════════════════════════════════════════════════
class HubPosition:
    def __init__(self):
        self.x = self.y = self.heading = 0.0
        self._prev_dist = 0

    def reset(self, x=0.0, y=0.0):
        self.x = x; self.y = y; self._prev_dist = 0

    def update(self, dist_mm, heading_deg, heading_offset_deg=0.0):
        """
        heading_offset_deg: correction added to the raw hub heading before
        computing dead-reckoning movement.

        The heading is snapped to the nearest 90 deg multiple before computing
        X/Y deltas so that straight moves always stay on a pure axis (no drift
        from tiny IMU heading errors during a straight run).
        The raw corrected heading is still stored in self.heading for drawing.
        """
        raw_heading = float(heading_deg)
        corrected   = raw_heading + heading_offset_deg
        self.heading = corrected          # full precision for arrow drawing

        delta = dist_mm - self._prev_dist
        self._prev_dist = dist_mm
        if abs(delta) > 500: delta = 0

        # Snap to nearest 90 deg so movement is always axis-aligned
        snapped = round(corrected / 90.0) * 90
        rad = math.radians(snapped)
        self.x += delta * math.sin(rad)
        self.y -= delta * math.cos(rad)

# ════════════════════════════════════════════════════════════════
#  CAMPO  (236 x 115 cm en mm)
# ════════════════════════════════════════════════════════════════
FIELD_W_MM = 2360.0
FIELD_H_MM = 1150.0

# ════════════════════════════════════════════════════════════════
#  EXPORTAR XLSX
# ════════════════════════════════════════════════════════════════
def export_xlsx(waypoints, path, input_log=None):
    wb   = openpyxl.Workbook()
    thin = Side(style="thin", color="888888")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def H(ws, r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=True, name="Arial", color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="0F3460")
        cell.alignment = Alignment(horizontal="center"); cell.border = brd

    def D(ws, r, c, v, bg=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = Alignment(horizontal="center")
        cell.border = brd; cell.font = Font(name="Arial", size=10)
        if bg: cell.fill = PatternFill("solid", start_color=bg.lstrip("#"))

    ws1 = wb.active; ws1.title = "Waypoints"
    for c, h in enumerate(["#","X (mm)","Y (mm)","X (cm)","Y (cm)","Color","HEX","Nota"], 1):
        H(ws1, 1, c, h)

    prev = None; acum = 0
    for i, wp in enumerate(waypoints, 1):
        ci = wp.get("color","#888888")
        D(ws1,i+1,1,i)
        D(ws1,i+1,2,round(wp["x_mm"])); D(ws1,i+1,3,round(wp["y_mm"]))
        D(ws1,i+1,4,round(wp["x_mm"]/10)); D(ws1,i+1,5,round(wp["y_mm"]/10))
        D(ws1,i+1,6,wp.get("label",""), bg=ci)
        D(ws1,i+1,7,ci)
        D(ws1,i+1,8,wp.get("note",""))
        prev = wp

    for w, c in zip([5,12,12,10,10,16,12,20], range(1,9)):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    if input_log:
        ws2 = wb.create_sheet("Inputs Replay")
        for c, h in enumerate(["#","Time (ms)","Speed","Turn","Motor","Motor Val","Turn Cmd"], 1):
            H(ws2, 1, c, h)
        for i, cmd in enumerate(input_log, 1):
            D(ws2, i+1, 1, i)
            D(ws2, i+1, 2, cmd["t_ms"])
            D(ws2, i+1, 3, cmd["speed"])
            D(ws2, i+1, 4, cmd["turn"])
            D(ws2, i+1, 5, cmd["mc"])
            D(ws2, i+1, 6, cmd["mv"])
            D(ws2, i+1, 7, cmd["tc"])
        for w, c in zip([5, 12, 10, 10, 10, 12, 12], range(1, 8)):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    wb.save(path)

# ════════════════════════════════════════════════════════════════
#  UI COLORS
# ════════════════════════════════════════════════════════════════
BG   = "#0F0F0F"
CARD = "#1A1A1A"
TEXT = "#FFFFFF"
DIM  = "#555555"
CYAN = "#00D4FF"
GRN  = "#00FF88"
AMB  = "#FFB800"
RED  = "#FF4444"
PRP  = "#BB86FC"
ORG  = "#FF9500"

# ════════════════════════════════════════════════════════════════
#  APP
# ════════════════════════════════════════════════════════════════
class WROApp:
    MAP_IMG  = "mapa_wro2026.jpg"

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("WRO 2026 — Control + Mapa")
        self.root.configure(bg=BG)
        self.root.geometry("1300x780")
        self.root.minsize(1000, 650)

        self._pub  = BLEPublisher()
        self._hub  = HubPosition()

        # ── HEADING OFFSET CALIBRATION ───────────────────────────
        # Positive value rotates the robot arrow CW on screen.
        # Adjust until "forward on screen" matches real robot forward.
        self._heading_offset = tk.IntVar(value=0)

        # Estado BLE
        self._hub_online  = False
        self._last_rgb    = (0, 0, 0)
        self._last_hex    = "#000000"

        # Control
        self._keys      = set()
        self._btn       = None
        self._last_cmd  = (0,0,0,0,0)
        self._mheld     = 0
        self._mdir      = 0
        self.spd        = tk.IntVar(value=300)

        # Mapa
        self._map_img_orig  = None
        self._map_photo     = None
        self._map_w = self._map_h = 1

        # Origen del robot en el canvas (pixels)
        self._origin_px = None
        self._robot_px  = None

        # Waypoints
        self._waypoints: list[dict] = []
        self._wp_selected = None

        # Modo de click
        self._click_mode = tk.StringVar(value="origin")

        # Color activo para waypoints
        self._active_color = "#FFD700"
        self._active_label = tk.StringVar(value="Punto")

        # Trail
        self._trail: deque = deque(maxlen=2000)

        # Input recording
        self._input_log: list[dict] = []
        self._recording   = False
        self._record_t0   = 0.0

        self._build()
        self._load_map()
        self._bind_keys()
        self._start_ble_rx()
        self._cmd_loop()

    # ─────────────────────────────────────────────────────────────
    def _build(self):
        # Toolbar
        tb = tk.Frame(self.root, bg="#111")
        tb.pack(fill="x")
        tk.Label(tb, text="WRO 2026 — Mosaic Masters",
                 bg="#111", fg=TEXT, font=("Helvetica",12,"bold")).pack(side="left",padx=12,pady=7)
        self.ble_lbl = tk.Label(tb, text="● Sin señal", bg="#111", fg=DIM,
                                font=("Helvetica",10))
        self.ble_lbl.pack(side="right", padx=12)
        self.color_indicator = tk.Canvas(tb, bg="#111", width=22, height=22,
                                         highlightthickness=0)
        self.color_indicator.pack(side="right", padx=2)
        self.color_indicator.create_oval(2,2,20,20, fill="#333", outline="#555", tags="dot")
        self.hex_lbl = tk.Label(tb, text="Sin señal", bg="#111", fg=DIM,
                                font=("Helvetica",10,"bold"))
        self.hex_lbl.pack(side="right", padx=4)
        tk.Label(tb, text="Color:", bg="#111", fg=DIM,
                 font=("Helvetica",9)).pack(side="right", padx=4)
        tk.Frame(self.root, bg="#252525", height=1).pack(fill="x")

        # Body
        body = tk.Frame(self.root, bg=BG)
        body.pack(fill="both", expand=True)

        # ── MAPA ──────────────────────────────────────────────────
        mf = tk.Frame(body, bg="#000")
        mf.pack(side="left", fill="both", expand=True)

        modebar = tk.Frame(mf, bg="#111")
        modebar.pack(fill="x")
        tk.Label(modebar, text="Modo click:", bg="#111", fg=DIM,
                 font=("Helvetica",9)).pack(side="left", padx=8, pady=4)
        for text, val, color in [
            ("📍 Colocar robot",  "origin",   AMB),
            ("🗺 Waypoint",       "waypoint",  CYAN),
        ]:
            tk.Radiobutton(modebar, text=text, variable=self._click_mode, value=val,
                           bg="#111", fg=color, selectcolor="#222",
                           activebackground="#111", activeforeground=color,
                           font=("Helvetica",9), indicatoron=True,
                           command=self._on_mode_change).pack(side="left", padx=6)

        self.mode_hint = tk.Label(modebar, text="", bg="#111", fg=AMB,
                                  font=("Helvetica",8,"italic"))
        self.mode_hint.pack(side="right", padx=10)

        self.cv = tk.Canvas(mf, bg="#1a1a1a", cursor="crosshair",
                            highlightthickness=0)
        self.cv.pack(fill="both", expand=True)
        self.cv.bind("<Configure>",   self._on_canvas_resize)
        self.cv.bind("<Button-1>",    self._on_canvas_click)
        self.cv.bind("<Button-3>",    self._on_canvas_right_click)

        self.pos_lbl = tk.Label(mf, text="X=0mm  Y=0mm  Hdg=0°",
                                bg="#111", fg=CYAN, font=("Courier",9))
        self.pos_lbl.pack(fill="x")

        # ── PANEL DERECHO ──────────────────────────────────────────
        pf = tk.Frame(body, bg=BG, width=280)
        pf.pack(side="right", fill="y", padx=(0,0))
        pf.pack_propagate(False)

        c2 = tk.Canvas(pf, bg=BG, highlightthickness=0)
        sb = tk.Scrollbar(pf, orient="vertical", command=c2.yview)
        c2.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        c2.pack(side="left", fill="both", expand=True)
        inn = tk.Frame(c2, bg=BG, width=265)
        c2.create_window((0,0), window=inn, anchor="nw", width=265)
        inn.bind("<Configure>", lambda e: c2.configure(scrollregion=c2.bbox("all")))

        def sep(): tk.Frame(inn, bg="#252525", height=1).pack(fill="x", pady=6)
        def lbl(t): tk.Label(inn, text=t, bg=BG, fg=DIM,
                             font=("Helvetica",8,"bold")).pack(anchor="w",padx=8,pady=(4,2))

        # ── D-PAD ────────────────────────────────────────────────
        lbl("RUEDAS  (WASD / flechas)")
        pad = tk.Frame(inn, bg=BG); pad.pack(pady=2)
        bkw = dict(width=4,height=2,relief="flat",cursor="hand2",
                   font=("Helvetica",18),bg=CARD,fg=TEXT,
                   activebackground="#1E3A4A",activeforeground=CYAN)

        tk.Label(pad,bg=BG,width=4,height=2).grid(row=0,column=0,padx=2,pady=2)
        self.b_fwd=tk.Button(pad,text="▲",**bkw); self.b_fwd.grid(row=0,column=1,padx=2,pady=2)
        tk.Label(pad,bg=BG,width=4,height=2).grid(row=0,column=2,padx=2,pady=2)

        self.b_left=tk.Button(pad,text="◀",**bkw); self.b_left.grid(row=1,column=0,padx=2,pady=2)
        tk.Button(pad,text="⏹",width=4,height=2,relief="flat",cursor="hand2",
                  font=("Helvetica",18),bg="#1E1E1E",fg=RED,
                  activebackground="#2A2A2A",activeforeground=RED,
                  command=self._estop).grid(row=1,column=1,padx=2,pady=2)
        self.b_right=tk.Button(pad,text="▶",**bkw); self.b_right.grid(row=1,column=2,padx=2,pady=2)

        tk.Label(pad,bg=BG,width=4,height=2).grid(row=2,column=0,padx=2,pady=2)
        self.b_back=tk.Button(pad,text="▼",**bkw); self.b_back.grid(row=2,column=1,padx=2,pady=2)
        tk.Label(pad,bg=BG,width=4,height=2).grid(row=2,column=2,padx=2,pady=2)

        for btn, act in [(self.b_fwd,"fwd"),(self.b_back,"back"),
                         (self.b_left,"left"),(self.b_right,"right")]:
            btn.bind("<ButtonPress-1>",   lambda e,a=act: self._bp(a))
            btn.bind("<ButtonRelease-1>", lambda e: self._br())

        sf = tk.Frame(inn,bg=BG); sf.pack(fill="x",padx=8,pady=3)
        tk.Label(sf,text="Vel:",bg=BG,fg=DIM,font=("Helvetica",9)).pack(side="left")
        tk.Scale(sf,variable=self.spd,from_=50,to=600,orient="horizontal",
                 bg=BG,fg=TEXT,troughcolor="#2A2A2A",highlightthickness=0,
                 sliderrelief="flat",activebackground=CYAN,length=160).pack(side="right")

        self.cmd_lbl = tk.Label(inn,text="spd=0  trn=0",bg=BG,fg=DIM,font=("Courier",8))
        self.cmd_lbl.pack()

        sep()

        # ══════════════════════════════════════════════════════════
        #  HEADING OFFSET CALIBRATION BLOCK
        # ══════════════════════════════════════════════════════════
        lbl("🧭 CALIBRAR DIRECCIÓN DEL ROBOT")

        cal_info = tk.Label(inn,
            text="Ajusta hasta que la flecha en pantalla\napunte igual que el robot real.",
            bg=BG, fg=DIM, font=("Helvetica",7,"italic"), justify="left")
        cal_info.pack(anchor="w", padx=10, pady=(0,4))

        # Compass rose canvas — shows robot arrow live
        self._compass_canvas = tk.Canvas(inn, bg=CARD, width=80, height=80,
                                         highlightthickness=1,
                                         highlightbackground="#333")
        self._compass_canvas.pack(pady=4)
        self._draw_compass()

        # ±1° / ±15° / ±90° buttons in two rows
        row1 = tk.Frame(inn,bg=BG); row1.pack(fill="x",padx=8,pady=1)
        row2 = tk.Frame(inn,bg=BG); row2.pack(fill="x",padx=8,pady=1)

        bk = dict(relief="flat", cursor="hand2", font=("Helvetica",9),
                  bg=CARD, fg=ORG, activebackground="#2A2A2A",
                  activeforeground="#FFF", padx=4, pady=3)

        for delta, label in [(-90,"↺90°"),(-15,"↺15°"),(-1,"↺1°")]:
            tk.Button(row1, text=label, **bk,
                      command=lambda d=delta: self._adjust_offset(d)
                      ).pack(side="left", padx=2, expand=True, fill="x")
        for delta, label in [(1,"↻1°"),(15,"↻15°"),(90,"↻90°")]:
            tk.Button(row2, text=label, **bk,
                      command=lambda d=delta: self._adjust_offset(d)
                      ).pack(side="left", padx=2, expand=True, fill="x")

        # Slider for fine/coarse adjustment
        slrow = tk.Frame(inn,bg=BG); slrow.pack(fill="x",padx=8,pady=3)
        tk.Label(slrow,text="Offset:",bg=BG,fg=DIM,font=("Helvetica",8)).pack(side="left")
        tk.Scale(slrow, variable=self._heading_offset,
                 from_=-180, to=180, orient="horizontal",
                 bg=BG, fg=ORG, troughcolor="#2A2A2A", highlightthickness=0,
                 sliderrelief="flat", activebackground=ORG, length=145,
                 command=lambda _: self._on_offset_change()
                 ).pack(side="right")

        # Value label + Reset
        oval_row = tk.Frame(inn,bg=BG); oval_row.pack(fill="x",padx=8,pady=(0,4))
        self.offset_lbl = tk.Label(oval_row, text="0°", bg=BG, fg=ORG,
                                   font=("Courier",9,"bold"))
        self.offset_lbl.pack(side="left")
        tk.Button(oval_row, text="Reset 0°", relief="flat", cursor="hand2",
                  font=("Helvetica",8), bg="#2A1A00", fg=ORG,
                  activebackground="#3A2A00", padx=6, pady=2,
                  command=self._reset_offset).pack(side="right")

        # Quick presets
        preset_row = tk.Frame(inn,bg=BG); preset_row.pack(fill="x",padx=8,pady=(0,4))
        tk.Label(preset_row,text="Presets:",bg=BG,fg=DIM,font=("Helvetica",7)).pack(side="left")
        for val, label in [(0,"↑ Norte"),(90,"→ Este"),(180,"↓ Sur"),(-90,"← Oeste")]:
            tk.Button(preset_row, text=label, relief="flat", cursor="hand2",
                      font=("Helvetica",7), bg=CARD, fg=DIM,
                      activebackground="#252525", padx=3, pady=2,
                      command=lambda v=val: self._set_offset(v)
                      ).pack(side="left", padx=1)

        sep()

        # ── GIROS PRECISOS ────────────────────────────────────────
        lbl("GIROS PRECISOS (ejecutados en el hub)")
        trow1 = tk.Frame(inn,bg=BG); trow1.pack(padx=8,pady=2)
        trow2 = tk.Frame(inn,bg=BG); trow2.pack(padx=8,pady=2)

        tkw = dict(width=5,height=2,relief="flat",cursor="hand2",
                   font=("Helvetica",12),bg=CARD,fg=CYAN,
                   activebackground="#1E3A4A",activeforeground="#FFF")

        for deg in [-90, -45]:
            tk.Button(trow1, text=f"↺{abs(deg)}°", **tkw,
                      command=lambda d=deg: self._do_turn(d)).pack(side="left",padx=3)
        for deg in [45, 90]:
            tk.Button(trow1, text=f"↻{deg}°", **tkw,
                      command=lambda d=deg: self._do_turn(d)).pack(side="left",padx=3)

        for deg in [-180, 180]:
            label = f"↺180°" if deg < 0 else f"↻180°"
            tk.Button(trow2, text=label, **tkw,
                      command=lambda d=deg: self._do_turn(d)).pack(side="left",padx=3)

        self.turn_lbl = tk.Label(inn,text="",bg=BG,fg=CYAN,font=("Courier",8))
        self.turn_lbl.pack()

        sep()

        # ── MOTORES EXTRA ─────────────────────────────────────────
        lbl("MOTORES EXTRA")
        motors = [(1,"Garra Principal","A",PRP),
                  (2,"Agarrar Bloques","E","#FF9500"),
                  (3,"Expandir Garra", "C","#00E5CC")]
        for mid, name, port, color in motors:
            mrow = tk.Frame(inn,bg=BG); mrow.pack(fill="x",padx=8,pady=2)
            tk.Label(mrow,text=f"{name} ({port})",bg=BG,fg=TEXT,
                     font=("Helvetica",8)).pack(side="left")
            bf = tk.Frame(mrow,bg=BG); bf.pack(side="right")
            mkw = dict(width=3,height=1,relief="flat",cursor="hand2",
                       font=("Helvetica",13),bg=CARD,fg=color,
                       activebackground="#2A2A2A")
            bm=tk.Button(bf,text="−",**mkw); bm.pack(side="left",padx=1)
            bp=tk.Button(bf,text="+",**mkw); bp.pack(side="left",padx=1)
            bm.bind("<ButtonPress-1>",   lambda e,m=mid: self._mpress(m,-1))
            bm.bind("<ButtonRelease-1>", lambda e: self._mrelease())
            bp.bind("<ButtonPress-1>",   lambda e,m=mid: self._mpress(m,+1))
            bp.bind("<ButtonRelease-1>", lambda e: self._mrelease())

        sep()

        # ── WAYPOINTS ────────────────────────────────────────────
        lbl("WAYPOINTS")

        wpc = tk.Frame(inn,bg=BG); wpc.pack(fill="x",padx=8,pady=2)
        tk.Label(wpc,text="Etiqueta:",bg=BG,fg=DIM,font=("Helvetica",9)).pack(side="left")
        tk.Entry(wpc,textvariable=self._active_label,bg=CARD,fg=TEXT,
                 insertbackground=TEXT,font=("Helvetica",9),width=12,relief="flat"
                 ).pack(side="left",padx=4)

        colrow = tk.Frame(inn,bg=BG); colrow.pack(fill="x",padx=8,pady=2)
        tk.Label(colrow,text="Color:",bg=BG,fg=DIM,font=("Helvetica",9)).pack(side="left")
        self.color_btn_preview = tk.Canvas(colrow,width=24,height=24,
                                           bg=self._active_color,highlightthickness=1,
                                           highlightbackground="#555",cursor="hand2")
        self.color_btn_preview.pack(side="left",padx=4)
        self.color_btn_preview.bind("<Button-1>", self._pick_color)
        self.hex_entry_var = tk.StringVar(value=self._active_color)
        hex_e = tk.Entry(colrow,textvariable=self.hex_entry_var,bg=CARD,fg=TEXT,
                         insertbackground=TEXT,font=("Courier",9),width=9,relief="flat")
        hex_e.pack(side="left",padx=2)
        hex_e.bind("<Return>", self._apply_hex)
        tk.Button(colrow,text="OK",bg=CARD,fg=CYAN,relief="flat",cursor="hand2",
                  font=("Helvetica",8),padx=4,
                  command=self._apply_hex).pack(side="left")

        qrow = tk.Frame(inn,bg=BG); qrow.pack(fill="x",padx=8,pady=2)
        tk.Label(qrow,text="Rápidos:",bg=BG,fg=DIM,font=("Helvetica",8)).pack(side="left")
        for hex_c, name in [("#FFD700","Amarillo"),("#2255CC","Azul"),
                             ("#22AA44","Verde"),("#DDDDDD","Blanco"),
                             ("#FF4444","Rojo")]:
            dot = tk.Canvas(qrow,width=20,height=20,bg=hex_c,
                            highlightthickness=1,highlightbackground="#555",cursor="hand2")
            dot.pack(side="left",padx=2)
            dot.bind("<Button-1>", lambda e,c=hex_c,n=name: self._set_quick_color(c,n))

        sep()

        lbl("LISTA DE WAYPOINTS")
        self.wp_listbox = tk.Listbox(inn,bg=CARD,fg=TEXT,selectbackground="#0F3460",
                                     font=("Courier",8),height=8,relief="flat",
                                     activestyle="none",highlightthickness=0)
        self.wp_listbox.pack(fill="x",padx=8,pady=2)
        self.wp_listbox.bind("<<ListboxSelect>>", self._on_wp_select)

        wpbrow = tk.Frame(inn,bg=BG); wpbrow.pack(fill="x",padx=8,pady=2)
        bkwp = dict(relief="flat",cursor="hand2",font=("Helvetica",9),pady=4,padx=6)
        tk.Button(wpbrow,text="↑",bg=CARD,fg=TEXT,**bkwp,
                  command=self._wp_up).pack(side="left",padx=2)
        tk.Button(wpbrow,text="↓",bg=CARD,fg=TEXT,**bkwp,
                  command=self._wp_down).pack(side="left",padx=2)
        tk.Button(wpbrow,text="✕ Borrar",bg="#3A1A1A",fg=RED,**bkwp,
                  command=self._wp_delete).pack(side="left",padx=2)
        tk.Button(wpbrow,text="🗑 Limpiar",bg=CARD,fg=DIM,**bkwp,
                  command=self._wp_clear).pack(side="left",padx=2)

        sep()

        sep()

        # ── GRABACIÓN ────────────────────────────────────────────
        lbl("GRABACIÓN DE INPUTS")

        rec_row = tk.Frame(inn, bg=BG); rec_row.pack(fill="x", padx=8, pady=2)
        bkr = dict(relief="flat", cursor="hand2", font=("Helvetica",9), pady=4, padx=6)
        tk.Button(rec_row, text="● Grabar", bg="#3A1A1A", fg=RED, **bkr,
                  command=self._start_recording).pack(side="left", padx=2)
        tk.Button(rec_row, text="■ Parar",  bg=CARD,     fg=DIM, **bkr,
                  command=self._stop_recording).pack(side="left", padx=2)
        tk.Button(rec_row, text="▶ Replay", bg="#1A3A1A", fg=GRN, **bkr,
                  command=self._start_replay).pack(side="left", padx=2)

        rec_row2 = tk.Frame(inn, bg=BG); rec_row2.pack(fill="x", padx=8, pady=2)
        tk.Button(rec_row2, text="🗑 Limpiar", bg=CARD, fg=DIM, **bkr,
                  command=self._clear_log).pack(side="left", padx=2)
        self._rec_lbl = tk.Label(rec_row2, text="Sin grabación", bg=BG, fg=DIM,
                                 font=("Courier", 8))
        self._rec_lbl.pack(side="left", padx=8)

        sep()

        bm2 = dict(relief="flat",cursor="hand2",font=("Helvetica",10),pady=7)
        tk.Button(inn,text="💾  Exportar XLSX",bg="#1A3A1A",fg=GRN,
                  activebackground="#2A4A2A",**bm2,
                  command=self._export).pack(fill="x",padx=8,pady=2)

        sep()
        lbl("POSICIÓN REAL")
        pos_card = tk.Frame(inn,bg=CARD); pos_card.pack(fill="x",padx=8,pady=2)
        self.pos_card_lbl = tk.Label(pos_card,
            text="X=0 mm\nY=0 mm\nHeading=0°",
            bg=CARD,fg=CYAN,font=("Courier",9),justify="left")
        self.pos_card_lbl.pack(padx=10,pady=6,anchor="w")

    # ─────────────────────────────────────────────────────────────
    #  HEADING OFFSET HELPERS
    # ─────────────────────────────────────────────────────────────
    def _adjust_offset(self, delta: int):
        new_val = (self._heading_offset.get() + delta + 360) % 360
        if new_val > 180: new_val -= 360          # keep in [-180, 180]
        self._heading_offset.set(new_val)
        self._on_offset_change()

    def _set_offset(self, val: int):
        self._heading_offset.set(val)
        self._on_offset_change()

    def _reset_offset(self):
        self._heading_offset.set(0)
        self._on_offset_change()

    def _on_offset_change(self):
        v = self._heading_offset.get()
        sign = "+" if v >= 0 else ""
        self.offset_lbl.config(text=f"{sign}{v}°")
        # Resets dead-reckoning trail when user recalibrates so drift doesn't compound
        if self._origin_px is not None:
            self._hub.reset()
            self._trail.clear()
            self._trail.append((0.0, 0.0))
        self._draw_compass()
        self._redraw()

    def _draw_compass(self):
        """Small compass rose showing current robot forward direction with offset applied."""
        cc = self._compass_canvas
        cc.delete("all")
        cx = cy = 40; r = 32
        # Circle
        cc.create_oval(cx-r, cy-r, cx+r, cy+r, outline="#333", width=1)
        # Cardinal labels
        for ang, label in [(0,"N"),(90,"E"),(180,"S"),(270,"W")]:
            rad = math.radians(ang)
            lx = cx + (r+8)*math.sin(rad)
            ly = cy - (r+8)*math.cos(rad)
            cc.create_text(lx, ly, text=label, fill="#444", font=("Helvetica",6))
        # Robot arrow — heading=0 + offset
        off = self._heading_offset.get() if hasattr(self, '_heading_offset') else 0
        rad = math.radians(off)
        tx = cx + r*0.85*math.sin(rad)
        ty = cy - r*0.85*math.cos(rad)
        # Arrow shaft
        cc.create_line(cx, cy, tx, ty, fill=AMB, width=3, arrow="last",
                       arrowshape=(8,10,4))
        # Center dot
        cc.create_oval(cx-3,cy-3,cx+3,cy+3, fill="#FFF", outline="")
        # Offset text
        v = off; sign = "+" if v >= 0 else ""
        cc.create_text(cx, cy+r+10, text=f"{sign}{v}°",
                       fill=ORG, font=("Courier",7,"bold"))

    # ─────────────────────────────────────────────────────────────
    #  MAP LOADING
    # ─────────────────────────────────────────────────────────────
    def _load_map(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        img_path   = os.path.join(script_dir, self.MAP_IMG)
        if not os.path.exists(img_path):
            img_path = self.MAP_IMG
        try:
            self._map_img_orig = Image.open(img_path)
            print(f"[Map] Loaded {self._map_img_orig.size}")
        except Exception as e:
            print(f"[Map] Could not load {img_path}: {e}")
            self._map_img_orig = None

    def _on_canvas_resize(self, event=None):
        self._render_map()

    def _render_map(self):
        w = self.cv.winfo_width()
        h = self.cv.winfo_height()
        if w < 10 or h < 10: return

        if self._map_img_orig:
            img_w, img_h = self._map_img_orig.size
            scale = min(w / img_w, h / img_h)
            nw = int(img_w * scale)
            nh = int(img_h * scale)
            img_resized = self._map_img_orig.resize((nw, nh), Image.LANCZOS)
            self._map_photo = ImageTk.PhotoImage(img_resized)
            self._map_w = nw
            self._map_h = nh
        else:
            self._map_photo = None
            self._map_w = w
            self._map_h = h

        self._redraw()

    # ─────────────────────────────────────────────────────────────
    #  COORDINATE CONVERSION
    # ─────────────────────────────────────────────────────────────
    def _mm_to_px(self, x_mm, y_mm):
        px = int(x_mm / FIELD_W_MM * self._map_w)
        py = int(y_mm / FIELD_H_MM * self._map_h)
        return px, py

    def _px_to_mm(self, px, py):
        x_mm = px / self._map_w * FIELD_W_MM
        y_mm = py / self._map_h * FIELD_H_MM
        return x_mm, y_mm

    # ─────────────────────────────────────────────────────────────
    #  CANVAS DRAWING
    # ─────────────────────────────────────────────────────────────
    def _redraw(self):
        cv = self.cv; cv.delete("all")
        w  = cv.winfo_width()
        h  = cv.winfo_height()

        cv.create_rectangle(0,0,w,h,fill="#000",outline="")

        if self._map_photo:
            ox = (w - self._map_w) // 2
            oy = (h - self._map_h) // 2
            cv.create_image(ox, oy, anchor="nw", image=self._map_photo)
        else:
            cv.create_text(w//2, h//2, text="mapa_wro2026.jpg no encontrado",
                           fill=DIM, font=("Helvetica",12))

        self._map_ox = (w - self._map_w) // 2
        self._map_oy = (h - self._map_h) // 2

        trail = list(self._trail)
        for i in range(1, len(trail)):
            x1,y1 = self._world_to_canvas(*trail[i-1])
            x2,y2 = self._world_to_canvas(*trail[i])
            cv.create_line(x1,y1,x2,y2, fill="#00AAAA", width=2)

        for i, wp in enumerate(self._waypoints):
            cx,cy = self._wp_canvas(wp)
            sel   = (i == self._wp_selected)
            r     = 9 if sel else 7
            col   = wp.get("color","#FFD700")
            cv.create_oval(cx-r,cy-r,cx+r,cy+r,
                           fill=col, outline="#FFF" if sel else "#AAA",
                           width=3 if sel else 1)
            cv.create_text(cx, cy+14, text=f"{i+1}:{wp.get('label','')}",
                           fill="#FFF", font=("Helvetica",7,"bold"),
                           tags="wplabel")

        for i in range(1, len(self._waypoints)):
            x1,y1 = self._wp_canvas(self._waypoints[i-1])
            x2,y2 = self._wp_canvas(self._waypoints[i])
            cv.create_line(x1,y1,x2,y2, fill=CYAN, width=1, dash=(5,3))

        if self._origin_px is not None:
            wx, wy = self._hub_world_pos()
            cx, cy = self._world_to_canvas(wx, wy)
            self._draw_robot(cx, cy, self._hub.heading)

    def _world_to_canvas(self, x_mm, y_mm):
        px, py = self._mm_to_px(x_mm, y_mm)
        return px + getattr(self,'_map_ox',0), py + getattr(self,'_map_oy',0)

    def _canvas_to_world(self, cx, cy):
        ox = getattr(self,'_map_ox',0)
        oy = getattr(self,'_map_oy',0)
        return self._px_to_mm(cx - ox, cy - oy)

    def _wp_canvas(self, wp):
        return self._world_to_canvas(wp["x_mm"], wp["y_mm"])

    def _hub_world_pos(self):
        if self._origin_px is None:
            return 0.0, 0.0
        ox_mm, oy_mm = self._canvas_to_world(*self._origin_px)
        return ox_mm + self._hub.x, oy_mm + self._hub.y

    def _draw_robot(self, cx, cy, heading_deg):
        """
        heading_deg here is already the corrected heading (raw + offset),
        stored in HubPosition.heading after update().
        Triangle tip points in the direction the robot is facing.
        """
        size = 14
        ang  = math.radians(heading_deg)
        tx = cx + size * math.sin(ang)
        ty = cy - size * math.cos(ang)
        al = ang + math.radians(145)
        ar = ang - math.radians(145)
        lx = cx + size*0.65*math.sin(al); ly = cy - size*0.65*math.cos(al)
        rx = cx + size*0.65*math.sin(ar); ry = cy - size*0.65*math.cos(ar)
        self.cv.create_polygon(tx,ty,lx,ly,rx,ry,
                               fill=AMB, outline="#FFF", width=2)
        self.cv.create_oval(cx-3,cy-3,cx+3,cy+3, fill="#FFF", outline="")

    # ─────────────────────────────────────────────────────────────
    #  CLICK HANDLERS
    # ─────────────────────────────────────────────────────────────
    def _on_mode_change(self):
        hints = {
            "origin":   "Haz click en el mapa para colocar el robot",
            "waypoint": "Haz click en el mapa para agregar un waypoint",
        }
        self.mode_hint.config(text=hints.get(self._click_mode.get(),""))

    def _on_canvas_click(self, event):
        mode = self._click_mode.get()
        cx, cy = event.x, event.y
        x_mm, y_mm = self._canvas_to_world(cx, cy)

        if mode == "origin":
            self._origin_px = (cx, cy)
            self._hub.reset()
            self._trail.clear()
            self._trail.append((0.0, 0.0))
            self._click_mode.set("waypoint")
            self._on_mode_change()
            self.mode_hint.config(text=f"✓ Robot colocado en ({round(x_mm/10)}cm, {round(y_mm/10)}cm)")

        elif mode == "waypoint":
            wp = {
                "x_mm":   x_mm,
                "y_mm":   y_mm,
                "color":  self._active_color,
                "label":  self._active_label.get(),
                "note":   "",
            }
            self._waypoints.append(wp)
            self._wp_selected = len(self._waypoints) - 1
            self._refresh_wp_list()

        self._redraw()

    def _on_canvas_right_click(self, event):
        for i, wp in enumerate(self._waypoints):
            cx, cy = self._wp_canvas(wp)
            if math.hypot(event.x - cx, event.y - cy) < 12:
                self._waypoints.pop(i)
                if self._wp_selected == i:
                    self._wp_selected = None
                self._refresh_wp_list()
                self._redraw()
                return

    # ─────────────────────────────────────────────────────────────
    #  COLOR PICKER
    # ─────────────────────────────────────────────────────────────
    def _pick_color(self, event=None):
        result = colorchooser.askcolor(color=self._active_color,
                                       title="Elegir color del waypoint")
        if result and result[1]:
            self._set_color(result[1])

    def _apply_hex(self, event=None):
        val = self.hex_entry_var.get().strip()
        if not val.startswith("#"): val = "#" + val
        try:
            self.root.winfo_rgb(val)
            self._set_color(val)
        except Exception:
            self.hex_entry_var.set(self._active_color)

    def _set_color(self, hex_c: str):
        self._active_color = hex_c
        self.hex_entry_var.set(hex_c)
        self.color_btn_preview.config(bg=hex_c)
        if self._wp_selected is not None and self._wp_selected < len(self._waypoints):
            self._waypoints[self._wp_selected]["color"] = hex_c
            self._refresh_wp_list()
            self._redraw()

    def _set_quick_color(self, hex_c, name):
        self._active_label.set(name)
        self._set_color(hex_c)

    # ─────────────────────────────────────────────────────────────
    #  WAYPOINT LIST
    # ─────────────────────────────────────────────────────────────
    def _refresh_wp_list(self):
        self.wp_listbox.delete(0, tk.END)
        for i, wp in enumerate(self._waypoints):
            text = f"{i+1:2d}. {wp.get('label',''):10s} ({round(wp['x_mm']/10):4d},{round(wp['y_mm']/10):4d})cm  {wp.get('color','')}"
            self.wp_listbox.insert(tk.END, text)
            self.wp_listbox.itemconfig(i, fg=wp.get("color","#FFF"),
                                       background="#1A1A1A")
        if self._wp_selected is not None:
            self.wp_listbox.selection_set(self._wp_selected)
            self.wp_listbox.see(self._wp_selected)

    def _on_wp_select(self, event):
        sel = self.wp_listbox.curselection()
        if sel:
            self._wp_selected = sel[0]
            wp = self._waypoints[self._wp_selected]
            self._set_color(wp.get("color", self._active_color))
            self._active_label.set(wp.get("label",""))
            self._redraw()

    def _wp_up(self):
        i = self._wp_selected
        if i and i > 0:
            self._waypoints[i], self._waypoints[i-1] = self._waypoints[i-1], self._waypoints[i]
            self._wp_selected = i-1
            self._refresh_wp_list(); self._redraw()

    def _wp_down(self):
        i = self._wp_selected
        if i is not None and i < len(self._waypoints)-1:
            self._waypoints[i], self._waypoints[i+1] = self._waypoints[i+1], self._waypoints[i]
            self._wp_selected = i+1
            self._refresh_wp_list(); self._redraw()

    def _wp_delete(self):
        i = self._wp_selected
        if i is not None and i < len(self._waypoints):
            self._waypoints.pop(i)
            self._wp_selected = min(i, len(self._waypoints)-1) if self._waypoints else None
            self._refresh_wp_list(); self._redraw()

    def _wp_clear(self):
        if messagebox.askyesno("Limpiar", "¿Borrar todos los waypoints?"):
            self._waypoints.clear(); self._wp_selected = None
            self._refresh_wp_list(); self._redraw()

    # ─────────────────────────────────────────────────────────────
    #  EXPORT
    # ─────────────────────────────────────────────────────────────
    def _export(self):
        if not self._waypoints:
            messagebox.showinfo("Sin datos","Agrega waypoints primero."); return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile="wro2026_ruta.xlsx")
        if not path: return
        try:
            export_xlsx(self._waypoints, path, input_log=self._input_log or None)
            n_cmds = len(self._input_log)
            msg = f"{len(self._waypoints)} waypoints"
            if n_cmds:
                msg += f"\n{n_cmds} inputs grabados"
            messagebox.showinfo("✓ Guardado", f"{msg}\n{path}")
        except Exception as e:
            messagebox.showerror("Error",str(e))

    # ─────────────────────────────────────────────────────────────
    #  TECLADO
    # ─────────────────────────────────────────────────────────────
    def _bind_keys(self):
        self.root.bind("<KeyPress>",   lambda e: self._keys.add(e.keysym.lower()))
        self.root.bind("<KeyRelease>", lambda e: self._keys.discard(e.keysym.lower()))
        self.root.focus_set()

    def _bp(self, a):
        self._btn = a
        m={"fwd":self.b_fwd,"back":self.b_back,"left":self.b_left,"right":self.b_right}
        b=m.get(a)
        if b: b.config(bg="#1E3A4A",fg=CYAN)

    def _br(self):
        m={"fwd":self.b_fwd,"back":self.b_back,"left":self.b_left,"right":self.b_right}
        b=m.get(self._btn)
        if b: b.config(bg=CARD,fg=TEXT)
        self._btn=None

    def _estop(self):
        self._btn=None; self._keys.clear()
        self._mheld=0; self._mdir=0
        self._pub.send(0,0,0,0,0)
        self.cmd_lbl.config(text="spd=0  trn=0")

    def _mpress(self, mid, direction):
        self._mheld=mid; self._mdir=direction

    def _mrelease(self):
        self._mheld=0; self._mdir=0

    # ─────────────────────────────────────────────────────────────
    #  GIRO PRECISO
    # ─────────────────────────────────────────────────────────────
    def _do_turn(self, degrees: int):
        if self._recording:
            t_ms = int((time.monotonic() - self._record_t0) * 1000)
            self._input_log.append({"t_ms": t_ms,       "speed": 0, "turn": 0, "mc": 0, "mv": 0, "tc": degrees})
            self._input_log.append({"t_ms": t_ms + 500, "speed": 0, "turn": 0, "mc": 0, "mv": 0, "tc": 0})
            self._rec_lbl.config(text=f"● REC  {len(self._input_log)} cmds", fg=RED)
        self._pub.send(0, 0, 0, 0, degrees)
        sign = "↺" if degrees < 0 else "↻"
        self.turn_lbl.config(text=f"Girando {sign} {abs(degrees)}°...", fg=CYAN)
        self.root.after(500, lambda: self._pub.send(0, 0, 0, 0, 0))
        self.root.after(2000, lambda: self.turn_lbl.config(text=""))

    # ─────────────────────────────────────────────────────────────
    #  GRABACIÓN / REPLAY
    # ─────────────────────────────────────────────────────────────
    def _start_recording(self):
        self._input_log.clear()
        self._recording  = True
        self._record_t0  = time.monotonic()
        self._rec_lbl.config(text="● REC  0 cmds", fg=RED)

    def _stop_recording(self):
        self._recording = False
        n = len(self._input_log)
        self._rec_lbl.config(text=f"Grabado: {n} cmds", fg=GRN if n else DIM)

    def _clear_log(self):
        self._input_log.clear()
        self._recording = False
        self._rec_lbl.config(text="Sin grabación", fg=DIM)

    def _start_replay(self):
        if not self._input_log:
            return
        self._rec_lbl.config(text="▶ Replay...", fg=CYAN)
        for entry in self._input_log:
            self.root.after(
                entry["t_ms"],
                lambda e=entry: self._pub.send(e["speed"], e["turn"], e["mc"], e["mv"], e["tc"])
            )
        last_t = self._input_log[-1]["t_ms"]
        self.root.after(last_t + 300, lambda: self._pub.send(0, 0, 0, 0, 0))
        self.root.after(last_t + 400, lambda: self._rec_lbl.config(text="Replay listo", fg=GRN))

    # ─────────────────────────────────────────────────────────────
    #  LOOP DE CONTROL (50ms)
    # ─────────────────────────────────────────────────────────────
    def _cmd_loop(self):
        k=self._keys; b=self._btn; s=self.spd.get()
        fwd  ="w" in k or "up"    in k or b=="fwd"
        back ="s" in k or "down"  in k or b=="back"
        left ="a" in k or "left"  in k or b=="left"
        right="d" in k or "right" in k or b=="right"
        spd = s if fwd else (-s if back else 0)
        trn = -TURN_RATE if right else (TURN_RATE if left else 0)
        mc  = self._mheld
        mv  = MOTOR_SPD * self._mdir if mc else 0
        cmd = (spd, trn, mc, mv, 0)
        if cmd != self._last_cmd:
            self._pub.send(*cmd)
            self._last_cmd = cmd
            if self._recording:
                t_ms = int((time.monotonic() - self._record_t0) * 1000)
                self._input_log.append({"t_ms": t_ms, "speed": spd, "turn": trn, "mc": mc, "mv": mv, "tc": 0})
                self._rec_lbl.config(text=f"● REC  {len(self._input_log)} cmds", fg=RED)
            moving = spd!=0 or trn!=0
            self.ble_lbl.config(
                text=f"● {'Activo' if moving else 'Conectado ✓'}",
                fg=GRN if moving else CYAN)
            self.cmd_lbl.config(text=f"spd={spd:+4d}  trn={trn:+4d}")
        self.root.after(50, self._cmd_loop)

    # ─────────────────────────────────────────────────────────────
    #  BLE RX
    # ─────────────────────────────────────────────────────────────
    def _start_ble_rx(self):
        threading.Thread(target=lambda: asyncio.run(self._scan()), daemon=True).start()

    async def _scan(self):
        def on_adv(device, adv):
            mfr = adv.manufacturer_data
            if PYBRICKS_MFR_ID not in mfr: return
            vals = decode_hub(mfr[PYBRICKS_MFR_ID])
            if vals is None: return
            dist_mm  = int(vals[0])
            heading  = int(vals[1])
            r, g, b  = int(vals[2]), int(vals[3]), int(vals[4])
            hex_c    = f"#{r:02X}{g:02X}{b:02X}"
            # ── Apply heading offset to dead-reckoning ──────────
            offset = self._heading_offset.get()
            self._hub.update(dist_mm, heading, heading_offset_deg=offset)
            wx, wy = self._hub_world_pos()
            self._trail.append((wx, wy))
            self._last_rgb = (r,g,b)
            self._last_hex = hex_c
            self._hub_online = True
            # Pass corrected heading to UI
            self.root.after(0, self._on_hub, self._hub.heading, hex_c, r, g, b)

        try:
            sc = BleakScanner(detection_callback=on_adv, scanning_mode="passive")
        except Exception:
            sc = BleakScanner(detection_callback=on_adv)
        async with sc:
            while True:
                await asyncio.sleep(1)

    # Nombres para los RGB exactos que envía spikemapper.py
    _CNAMES = {
        (220,  50,  50): "Rojo",
        (220, 200,  50): "Amarillo",
        ( 50, 200,  50): "Verde",
        ( 50,  50, 220): "Azul",
        (220, 220, 220): "Blanco",
        ( 40,  40,  40): "Sin color",
    }

    def _on_hub(self, heading, hex_c, r, g, b):
        wx, wy = self._hub_world_pos()
        self.pos_lbl.config(
            text=f"X={round(wx)}mm  Y={round(wy)}mm  Hdg={round(heading)}°")
        self.pos_card_lbl.config(
            text=f"X={round(wx)} mm\nY={round(wy)} mm\nHeading={round(heading)}°")
        self.ble_lbl.config(text="● Conectado ✓", fg=GRN)

        cname   = self._CNAMES.get((r, g, b), hex_c)
        is_dark = r + g + b < 80
        lbl_fg  = TEXT if is_dark else hex_c
        outline = "#AAA" if is_dark else "#888"
        self.hex_lbl.config(text=f"{cname}", fg=lbl_fg)
        self.color_indicator.delete("dot")
        self.color_indicator.create_oval(2, 2, 20, 20, fill=hex_c,
                                         outline=outline, tags="dot")
        self._redraw()

# ════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    app  = WROApp(root)
    root.protocol("WM_DELETE_WINDOW", lambda: (app._pub.stop(), root.destroy()))
    root.mainloop()

if __name__ == "__main__":
    main()